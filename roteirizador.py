import streamlit as st
import pandas as pd
import numpy as np
import folium
from folium.plugins import MarkerCluster, HeatMap
from streamlit_folium import st_folium
import io
import zipfile
import html
import re
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import time
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor
import os
import base64
import math 

# === MÓDULO DE INTELIGÊNCIA ARTIFICIAL E GRÁFICOS ===
try:
    from ortools.constraint_solver import routing_enums_pb2
    from ortools.constraint_solver import pywrapcp
except ImportError:
    st.error("🚨 Biblioteca 'ortools' não encontrada. Instale executando: pip install ortools")
    st.stop()

try:
    import plotly.express as px
except ImportError:
    st.error("🚨 Biblioteca 'plotly' não encontrada. Instale executando: pip install plotly")
    st.stop()

try:
    from sklearn.cluster import KMeans as SKKMeans
    from sklearn.cluster import DBSCAN
except ImportError:
    st.error("🚨 Biblioteca 'scikit-learn' não encontrada. Instale executando: pip install scikit-learn")
    st.stop()

# ==========================================
# 1. CONFIGURAÇÕES E CONSTANTES GERAIS
# ==========================================
STATUS_PADRAO = ['EM LEVANTAMENTO', '0', 'SEM INFORMAÇÕES', 'SEM INFORMACOES', 'CORREÇÃO DE LEVANTAMENTO', 'CORRECAO DE LEVANTAMENTO', 'PRÉ ANÁLISE', 'PRE ANALISE']
TIPOS_PRIORITARIOS = ["CCF", "DIF", "MGD", "MTP", "ASC", "SID"]
LOGO_PATH = "logo_nip-removebg-preview.png"

def get_retry_session(retries=4, backoff_factor=0.3):
    session = requests.Session()
    retry = Retry(total=retries, read=retries, connect=retries, backoff_factor=backoff_factor, status_forcelist=(429, 500, 502, 503, 504))
    adapter = HTTPAdapter(max_retries=retry, pool_connections=100, pool_maxsize=100)
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    return session

http_session = get_retry_session()

icon_page = LOGO_PATH if os.path.exists(LOGO_PATH) else "⚡"

st.set_page_config(
    page_title="Roteirizador NIP v1.1",
    page_icon=icon_page,
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==========================================
# 2. INJEÇÃO DE CSS E CACHE DE PERFORMANCE
# ==========================================
st.markdown("""
<style>
    header { display: none !important; }
    [data-testid="stHeader"] { display: none !important; }
    #MainMenu { display: none !important; }
    .stDeployButton { display: none !important; }
    [data-testid="stToolbar"] { display: none !important; }
    .viewerBadge_container { display: none !important; }
    [data-testid="manage-app-button"] { display: none !important; }
    footer { display: none !important; }
    .block-container { padding-top: 1.5rem !important; padding-bottom: 2rem !important; }
    .stSelectbox label, .stFileUploader label, .stRadio label, .stNumberInput label, .stMultiSelect label { font-size: 14px !important; font-weight: 700 !important; color: #0D256C !important; }
    .stepper-container { display: flex; justify-content: space-between; margin-top: 0.5rem; margin-bottom: 1.5rem; padding: 0.85rem 1.2rem; background: rgba(13, 37, 108, 0.04); border-radius: 10px; border: 1px solid rgba(13, 37, 108, 0.12); }
    .step-item { font-size: 13px; font-weight: 700; color: #6c757d; display: flex; align-items: center; gap: 6px; }
    .step-item.active { color: #0D256C; }
    .step-item.done { color: #55B929; }
    .metric-card { background: #ffffff; border: 1px solid #e2e8f0; border-radius: 12px; padding: 18px; box-shadow: 0 4px 8px rgba(0,0,0,0.04); display: flex; align-items: center; gap: 16px; margin-bottom: 12px; transition: transform 0.2s ease, box-shadow 0.2s ease; }
    .metric-card:hover { transform: translateY(-2px); box-shadow: 0 8px 16px rgba(0,0,0,0.08); }
    .metric-icon { font-size: 26px; padding: 12px; border-radius: 10px; display: flex; align-items: center; justify-content: center; }
    .metric-content .metric-title { font-size: 12px; font-weight: 700; color: #64748b; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 2px; }
    .metric-content .metric-value { font-size: 24px; font-weight: 800; color: #0D256C; }
    .profiling-box { background: rgba(85, 185, 41, 0.08); border-left: 4px solid #55B929; padding: 15px; border-radius: 6px; margin-bottom: 20px;}
    .brand-title { font-size: 28px; font-weight: 800; color: #0D256C; margin: 0; text-align: center; padding-bottom: 15px; }
</style>
""", unsafe_allow_html=True)

@st.cache_data(show_spinner=False)
def ler_planilha_cached(file_content):
    return pd.read_excel(io.BytesIO(file_content))

def formatar_moeda(valor):
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def formata_campo_html(val):
    val_str = str(val)
    if val_str.lower() == 'nan': return '-'
    if '|' in val_str:
        itens = val_str.split('|')
        if len(itens) > 1:
            lis = "".join([f"<li style='margin-bottom:3px;'><b>[{idx+1}]</b> {html.escape(i.strip())}</li>" for idx, i in enumerate(itens)])
            return f"<div style='max-height:95px; overflow-y:auto; border:1px solid #ccc; padding:6px; background:#fff; border-radius:4px;'><ul style='margin:0; padding-left:0px; list-style-type:none; font-size:11px; color:#333;'>{lis}</ul></div>"
    return html.escape(val_str)

def tentar_rerun():
    try:
        st.rerun()
    except AttributeError:
        st.experimental_rerun()

def limpar_roteirizador():
    st.session_state.roteamento_concluido = False
    st.session_state.vrp_status = "IDLE"
    st.session_state.vrp_state = {}
    st.session_state.df_routed = pd.DataFrame()
    st.session_state.bases_records = []
    st.session_state.tipo_periodo = "Dia"
    st.session_state.colunas_exibir = []
    st.session_state.col_prioridade = "TIPO NOTA"
    st.session_state.colunas_originais = []
    if 'bytes_zip_xl' in st.session_state: del st.session_state['bytes_zip_xl']
    if 'bytes_zip_kml' in st.session_state: del st.session_state['bytes_zip_kml']
    ler_planilha_cached.clear()
    tentar_rerun()

def normalize_cols(cols):
    new_cols = []
    for c in cols:
        c = str(c).strip().upper()
        c = re.sub(r'[ÁÀÂÃÄ]', 'A', c)
        c = re.sub(r'[ÉÈÊË]', 'E', c)
        c = re.sub(r'[ÍÌÎÏ]', 'I', c)
        c = re.sub(r'[ÓÒÔÕÖ]', 'O', c)
        c = re.sub(r'[ÚÙÛÜ]', 'U', c)
        c = re.sub(r'Ç', 'C', c)
        new_cols.append(c)
    return new_cols

def normalizar_municipios(series_mun):
    s = series_mun.astype(str).str.upper()
    s = s.str.replace(r'[ÁÀÂÃÄ]', 'A', regex=True)
    s = s.str.replace(r'[ÉÈÊË]', 'E', regex=True)
    s = s.str.replace(r'[ÍÌÎÏ]', 'I', regex=True)
    s = s.str.replace(r'[ÓÒÔÕÖ]', 'O', regex=True)
    s = s.str.replace(r'[ÚÙÛÜ]', 'U', regex=True)
    s = s.str.replace(r'Ç', 'C', regex=True)
    return s.str.split('-').str[0].str.strip()

def atualizar_status_via_df(df_principal, df_status, coluna_alvo):
    try:
        chave_nome = df_status.columns[0]
        df_status[chave_nome] = df_status[chave_nome].astype(str).str.strip()
        df_status_map = df_status.set_index(chave_nome)[coluna_alvo].to_dict()
        if 'PROTOCOLO' in df_principal.columns:
            df_principal['PROTOCOLO_STR'] = df_principal['PROTOCOLO'].astype(str).str.strip()
            df_principal['STATUS LIST'] = df_principal['PROTOCOLO_STR'].map(df_status_map).fillna(df_principal.get('STATUS LIST', 'SEM INFORMAÇÕES'))
            df_principal = df_principal.drop(columns=['PROTOCOLO_STR'])
            st.success(f"✅ Status Sincronizados: {len(df_status_map)} registros atualizados!")
        else:
            st.warning("⚠️ Coluna 'PROTOCOLO' não encontrada na base principal.")
    except Exception as e:
        st.error(f"Erro na sincronização: {e}")
    return df_principal

# ==========================================
# 3. TRATAMENTO DE SUPER PONTOS
# ==========================================
def fundir_super_pontos(df_tasks, raio_metros=5):
    if df_tasks.empty or 'LATITUDE' not in df_tasks.columns or 'LONGITUDE' not in df_tasks.columns:
        return df_tasks, 0
    df_valid = df_tasks.dropna(subset=['LATITUDE', 'LONGITUDE']).copy()
    if df_valid.empty: return df_tasks, 0

    coords = np.radians(df_valid[['LATITUDE', 'LONGITUDE']].values)
    eps_rad = raio_metros / 6371000.0
    
    db = DBSCAN(eps=eps_rad, min_samples=1, algorithm='ball_tree', metric='haversine').fit(coords)
    df_valid['CLUSTER_ID'] = db.labels_
    
    cluster_counts = df_valid['CLUSTER_ID'].value_counts()
    single_clusters = cluster_counts[cluster_counts == 1].index
    multi_clusters = cluster_counts[cluster_counts > 1].index
    
    agrupado = []
    
    if len(multi_clusters) > 0:
        df_multi = df_valid[df_valid['CLUSTER_ID'].isin(multi_clusters)]
        for c_id, group in df_multi.groupby('CLUSTER_ID'):
            row_base = group.iloc[0].copy()
            qtd = len(group)
            
            orig_rows = group.to_dict('records')
            row_base['_ORIGINAL_ROWS'] = orig_rows
            
            def safe_list_join(col_name):
                return " | ".join([str(x).strip() if pd.notna(x) else "-" for x in group[col_name]])
            
            if 'PROTOCOLO' in group.columns: row_base['PROTOCOLO'] = safe_list_join('PROTOCOLO')
            if 'CONTA CONTRATO' in group.columns: row_base['CONTA CONTRATO'] = safe_list_join('CONTA CONTRATO')
            if 'INSTALACAO' in group.columns: row_base['INSTALACAO'] = safe_list_join('INSTALACAO')
            if 'NOME' in group.columns: row_base['NOME'] = safe_list_join('NOME')
                
            row_base['LATITUDE'] = group['LATITUDE'].mean()
            row_base['LONGITUDE'] = group['LONGITUDE'].mean()
            row_base['SUPER_PONTO'] = f"SIM ({qtd} un.)"
            if 'PRIORIDADE' in group.columns and 'Sim' in group['PRIORIDADE'].values:
                row_base['PRIORIDADE'] = 'Sim'
                
            agrupado.append(row_base)
            
    df_final_multi = pd.DataFrame(agrupado)
    
    df_single = df_valid[df_valid['CLUSTER_ID'].isin(single_clusters)].copy()
    if not df_single.empty:
        df_single['SUPER_PONTO'] = "NÃO"
        dict_records = df_single.to_dict('records')
        df_single['_ORIGINAL_ROWS'] = [[r] for r in dict_records]
        
    df_final = pd.concat([df_single, df_final_multi], ignore_index=True).drop(columns=['CLUSTER_ID'], errors='ignore')
    
    df_nan = df_tasks[df_tasks['LATITUDE'].isna() | df_tasks['LONGITUDE'].isna()].copy()
    if not df_nan.empty:
        df_nan['SUPER_PONTO'] = "NÃO"
        dict_records_nan = df_nan.to_dict('records')
        df_nan['_ORIGINAL_ROWS'] = [[r] for r in dict_records_nan]
        df_final = pd.concat([df_final, df_nan], ignore_index=True)
        
    return df_final, len(df_tasks) - len(df_final)

# ==========================================
# 4. GEOCODING E MATEMÁTICA DE VETORES
# ==========================================
def geocode_endereco_nominatim(endereco, municipio):
    if pd.isna(endereco) or str(endereco).strip() == "" or pd.isna(municipio): 
        return np.nan, np.nan
    query = f"{str(endereco).strip()}, {str(municipio).strip()}, Maranhão, Brasil"
    url = "https://nominatim.openstreetmap.org/search"
    try:
        r = http_session.get(url, params={"q": query, "format": "json", "limit": 1}, headers={"User-Agent": "RoteirizadorNIP/1.0"}, timeout=5)
        if r.status_code == 200 and len(r.json()) > 0:
            return float(r.json()[0]['lat']), float(r.json()[0]['lon'])
    except Exception as e: 
        pass
    return np.nan, np.nan

def haversine_vectorized(lat1, lon1, lat2, lon2):
    R = 6371.0 
    lat1, lon1, lat2, lon2 = map(np.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = np.sin(dlat/2.0)**2 + np.cos(lat1) * np.cos(lat2) * np.sin(dlon/2.0)**2
    c = 2 * np.arcsin(np.sqrt(a))
    return R * c

def haversine_scalar(lat1, lon1, lat2, lon2):
    R = 6371.0
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat/2.0)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2.0)**2
    c = 2 * math.asin(math.sqrt(a))
    return R * c

def calcular_matriz_distancias_numpy(coords):
    R = 6371000.0
    lats = np.radians(coords[:, 0])
    lons = np.radians(coords[:, 1])
    dlat = lats[:, np.newaxis] - lats
    dlon = lons[:, np.newaxis] - lons
    a = np.sin(dlat / 2.0)**2 + np.cos(lats)[:, np.newaxis] * np.cos(lats) * np.sin(dlon / 2.0)**2
    c = 2 * np.arcsin(np.sqrt(a))
    return (R * c).astype(int)

def obter_matriz_osrm(coords, url_osrm_base):
    if len(coords) > 100 and 'project-osrm' in url_osrm_base: return None
    coords_str = ";".join([f"{lon:.6f},{lat:.6f}" for lat, lon in coords])
    url = f"{url_osrm_base}/table/v1/driving/{coords_str}?annotations=distance"
    try:
        r = http_session.get(url, timeout=5)
        if r.status_code == 200:
            data = r.json()
            if data.get('code') == 'Ok': return np.array(data['distances']).astype(int).tolist()
    except Exception as e: 
        pass
    return None

def resolver_tsp_ortools(lista_obras, base_lat, base_lon, url_osrm_base):
    if not lista_obras: return []
    coords_array = np.array([(base_lat, base_lon)] + [(r['LATITUDE'], r['LONGITUDE']) for r in lista_obras])
    
    distance_matrix = obter_matriz_osrm(coords_array, url_osrm_base)
    if distance_matrix is None:
        distance_matrix = calcular_matriz_distancias_numpy(coords_array).tolist()
        
    manager = pywrapcp.RoutingIndexManager(len(distance_matrix), 1, 0)
    routing = pywrapcp.RoutingModel(manager)

    def distance_callback(from_index, to_index):
        return distance_matrix[manager.IndexToNode(from_index)][manager.IndexToNode(to_index)]

    transit_callback_index = routing.RegisterTransitCallback(distance_callback)
    routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)
    
    penalty = 100000000 
    for node in range(1, len(distance_matrix)):
        routing.AddDisjunction([manager.NodeToIndex(node)], penalty)

    search_parameters = pywrapcp.DefaultRoutingSearchParameters()
    search_parameters.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
    search_parameters.local_search_metaheuristic = routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    search_parameters.time_limit.seconds = 2 
    
    solution = routing.SolveWithParameters(search_parameters)
    
    rota_atual = []
    if solution:
        index = routing.Start(0)
        while not routing.IsEnd(index):
            node_index = manager.IndexToNode(index)
            if node_index != 0:
                rota_atual.append(lista_obras[node_index - 1])
            index = solution.Value(routing.NextVar(index))
    return rota_atual

@st.cache_data(show_spinner=False)
def obter_coordenadas_municipio_cached(municipio):
    if not municipio or pd.isna(municipio) or str(municipio).strip() == "": return np.nan, np.nan
    try:
        url = f"https://nominatim.openstreetmap.org/search?q={str(municipio).strip()},+Maranhão,+Brasil&format=json&limit=1"
        r = http_session.get(url, headers={"User-Agent": "RoteirizadorNIP/1.0"}, timeout=5)
        if r.status_code == 200 and len(r.json()) > 0: return float(r.json()[0]['lat']), float(r.json()[0]['lon'])
    except: pass
    return np.nan, np.nan

def obter_rota_ruas(lat1, lon1, lat2, lon2, url_osrm_base, vel_fallback_kmh=30):
    try:
        url = f"{url_osrm_base}/route/v1/driving/{lon1:.6f},{lat1:.6f};{lon2:.6f},{lat2:.6f}?overview=full&geometries=geojson"
        r = http_session.get(url, timeout=4)
        if r.status_code == 200 and r.json().get('code') == 'Ok':
            return r.json()['routes'][0]['geometry']['coordinates'], r.json()['routes'][0]['duration']
    except Exception as e:
        pass
    
    coords = np.array([[lat1, lon1], [lat2, lon2]])
    dist_m = calcular_matriz_distancias_numpy(coords)[0][1]
    return [[lon1, lat1], [lon2, lat2]], (dist_m / 1000.0 / vel_fallback_kmh) * 3600

# ==========================================
# 5. MÓDULO DE EXPORTAÇÃO (EXCEL / KML)
# ==========================================
def identificar_icone_folium(row, colunas):
    tipo_str = str(row.get('TIPO LIGACAO', '')) + str(row.get('SERVICO', '')) + str(row.get('TIPO NOTA', ''))
    tipo_str = tipo_str.upper()
    if row.get('PROTOCOLO') == 'RETORNO_BASE': return 'home'
    if row.get('PROTOCOLO') == 'PAUSA_ALMOCO': return 'cutlery'
    if 'NOVA' in tipo_str or 'LIGACAO' in tipo_str or 'UNI' in tipo_str or 'UNR' in tipo_str: return 'bolt'
    if 'MANUT' in tipo_str or 'REPARO' in tipo_str: return 'wrench'
    if 'INSP' in tipo_str or 'VISTORIA' in tipo_str: return 'eye-open'
    return 'info-sign'

def gerar_excel_bytes(df, col_prioridade, colunas_originais=None):
    df_export = df.copy()
    if 'PROTOCOLO' in df_export.columns: 
        df_export = df_export[~df_export['PROTOCOLO'].isin(['RETORNO_BASE', 'PAUSA_ALMOCO'])]
        
    unpacked_rows = []
    for _, row in df_export.iterrows():
        if '_ORIGINAL_ROWS' in row and isinstance(row['_ORIGINAL_ROWS'], list):
            for orig in row['_ORIGINAL_ROWS']:
                new_row = orig.copy()
                for vrp_col in ['ORDEM', 'SEMANA', 'DIA', 'PERIODO', 'DISTANCIA_PONTO_ANTERIOR_KM', 'DISTANCIA_PROXIMO_PONTO_KM', 'TEMPO_VIAGEM_MINUTOS', 'HORA_INICIO', 'HORA_FIM', 'SUPER_PONTO', 'BASE_ATRIBUIDA', 'PRIORIDADE']:
                    if vrp_col in row:
                        new_row[vrp_col] = row[vrp_col]
                unpacked_rows.append(new_row)
        else:
            unpacked_rows.append(row.to_dict())
            
    df_export = pd.DataFrame(unpacked_rows)

    for col in ['ROTA_GEOMETRIA', 'STATUS LIST', 'INICIO AVARIA', 'STATUS ATUAL (LEVANTAMENTO)', 'DESCRICAO', '_HORA_INICIO_DT', '_HORA_FIM_DT', '_ORIGINAL_ROWS']:
        if col in df_export.columns: df_export = df_export.drop(columns=[col], errors='ignore')

    if colunas_originais:
        cols_atuais = df_export.columns.tolist()
        cols_originais_validas = [c for c in colunas_originais if c in cols_atuais]
        cols_novas_geradas = [c for c in cols_atuais if c not in colunas_originais]
        df_export = df_export[cols_originais_validas + cols_novas_geradas]
        
    buf_xl = io.BytesIO()
    with pd.ExcelWriter(buf_xl, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Roteiro')
        ws = writer.sheets['Roteiro']
        
        header_fill = PatternFill(start_color='0D256C', end_color='0D256C', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        col_types = {}
        for col_idx, col_name in enumerate(df_export.columns, 1):
            col_letter = get_column_letter(col_idx)
            col_name_upper = str(col_name).upper()
            if any(x in col_name_upper for x in ['NOME', 'CLIENTE', 'ENDEREÇO', 'INFORMAÇ']): ws.column_dimensions[col_letter].width = 45.0
            elif any(x in col_name_upper for x in ['PROTOCOLO', 'MUNICIPIO', 'BASE', 'LOCALIDADE']): ws.column_dimensions[col_letter].width = 25.0
            else: ws.column_dimensions[col_letter].width = 18.0
                
            if col_name_upper in ['ORDEM', 'SEMANA', 'DIA', 'PERIODO', 'DISTANCIA_PONTO_ANTERIOR_KM', 'DISTANCIA_PROXIMO_PONTO_KM', 'TEMPO_VIAGEM_MINUTOS', 'PRIORIDADE', 'HORA_INICIO', 'HORA_FIM']:
                col_types[col_idx] = center_align
            else:
                col_types[col_idx] = left_align

        red_font = Font(color="FF0000", bold=True)
        prio_idx = df_export.columns.get_loc('PRIORIDADE') + 1 if 'PRIORIDADE' in df_export.columns else None
        prio_target_idx = df_export.columns.get_loc(col_prioridade) + 1 if col_prioridade in df_export.columns else None

        for row_idx in range(2, len(df_export) + 2):
            for col_idx in range(1, len(df_export.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = col_types.get(col_idx, left_align)
            if prio_idx and ws.cell(row=row_idx, column=prio_idx).value == "Sim":
                ws.cell(row=row_idx, column=prio_idx).font = red_font
                if prio_target_idx and col_prioridade != "Nenhuma":
                    try: ws.cell(row=row_idx, column=prio_target_idx).font = red_font
                    except: pass
                    
        if 'SUPER_PONTO' in df_export.columns:
            try:
                idx_super = df_export.columns.get_loc('SUPER_PONTO') + 1
                yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                for row_idx in range(2, len(df_export) + 2):
                    val = str(ws.cell(row=row_idx, column=idx_super).value)
                    if val.startswith("SIM"):
                        for col_idx_f in range(1, len(df_export.columns) + 1):
                            ws.cell(row=row_idx, column=col_idx_f).fill = yellow_fill
            except: pass
                    
    return buf_xl.getvalue()

def gerar_excel_resumo_bytes(df):
    buf_xl = io.BytesIO()
    with pd.ExcelWriter(buf_xl, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Resumo')
        ws = writer.sheets['Resumo']
        
        header_fill = PatternFill(start_color='0D256C', end_color='0D256C', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            if str(col_name).upper() == 'LEVANTADOR':
                ws.column_dimensions[col_letter].width = 45.0
                col_align = left_align
            else:
                ws.column_dimensions[col_letter].width = 22.0
                col_align = center_align
            for row_idx in range(2, len(df) + 2): ws.cell(row=row_idx, column=col_idx).alignment = col_align
    return buf_xl.getvalue()

def gerar_kml_agrupado(df_rota, bases_records, doc_name, cols_exibir, lista_todas_bases=None, tipo_periodo="Dia"):
    if lista_todas_bases is None: lista_todas_bases = df_rota['BASE_ATRIBUIDA'].unique().tolist()
        
    kml_lines = [f'''<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
<Document>
  <name>{doc_name}</name>
  <Style id="linha-rota-contorno"><LineStyle><color>ff000000</color><width>8</width></LineStyle></Style>
  <Style id="icon-blue"><IconStyle><scale>1.1</scale><Icon><href>http://maps.google.com/mapfiles/kml/paddle/blu-blank.png</href></Icon><hotSpot x="32" xunits="pixels" y="64" yunits="insetPixels"/></IconStyle></Style>
  <Style id="icon-red"><IconStyle><scale>1.3</scale><Icon><href>http://maps.google.com/mapfiles/kml/paddle/red-blank.png</href></Icon><hotSpot x="32" xunits="pixels" y="64" yunits="insetPixels"/></IconStyle></Style>
  <Style id="icon-green"><IconStyle><scale>1.2</scale><Icon><href>https://maps.google.com/mapfiles/kml/shapes/homegardenbusiness.png</href></Icon></IconStyle></Style>
  <Style id="icon-yellow"><IconStyle><scale>1.3</scale><Icon><href>http://maps.google.com/mapfiles/kml/paddle/ylw-blank.png</href></Icon></IconStyle></Style>''']

    kml_cores = ['ff4b19e6', 'ffd4bc00', 'ffb5513f', 'ff889600', 'ff0098ff', 'ffb0279c', 'ff39dccd', 'ff631ee9', 'ff3bebff', 'ff485579']
    for idx, b_nome in enumerate(lista_todas_bases):
        cor_kml = kml_cores[idx % len(kml_cores)]
        nome_limpo = re.sub(r'[^A-Za-z0-9_]', '', str(b_nome))
        kml_lines.append(f'  <Style id="rota-centro-{nome_limpo}"><LineStyle><color>{cor_kml}</color><width>5</width></LineStyle></Style>')

    for base_nome in df_rota['BASE_ATRIBUIDA'].unique():
        df_base = df_rota[df_rota['BASE_ATRIBUIDA'] == base_nome]
        base_ref = next((b for b in bases_records if b['LEVANTADOR'] == base_nome), None)
        b_lat, b_lon = float(str(base_ref['LATITUDE']).replace(',','.')), float(str(base_ref['LONGITUDE']).replace(',','.'))
        res_nome = str(base_ref.get('RESIDENCIA', base_nome))
        nome_limpo_base = re.sub(r'[^A-Za-z0-9_]', '', str(base_nome))

        kml_lines.append(f'  <Folder>\n    <name>Levantador: {html.escape(str(base_nome))}</name>')
        kml_lines.append(f'    <Placemark><name>BASE: {html.escape(str(res_nome))}</name><styleUrl>#icon-green</styleUrl><Point><coordinates>{b_lon},{b_lat},0</coordinates></Point></Placemark>')

        for semana in df_base['SEMANA'].unique():
            df_semana = df_base[df_base['SEMANA'] == semana]
            kml_lines.append(f'    <Folder>\n      <name>Semana {semana}</name>')

            for dia in df_semana['DIA'].unique():
                df_dia = df_semana[df_semana['DIA'] == dia].copy().sort_values(by='ORDEM')
                kml_lines.append(f'      <Folder>\n        <name>Dia {dia}</name>')

                coords_linha_kml = []
                
                for r in df_dia.to_dict('records'):
                    lon, lat = str(r.get('LONGITUDE')).replace(',','.'), str(r.get('LATITUDE')).replace(',','.')
                    
                    if r.get('PROTOCOLO') in ['RETORNO_BASE', 'PAUSA_ALMOCO']:
                        geometria = r.get('ROTA_GEOMETRIA')
                        if isinstance(geometria, list):
                            coords_linha_kml.extend([f"          {pt_lon},{pt_lat},0" for pt_lon, pt_lat in geometria])
                        else:
                            coords_linha_kml.append(f"          {lon},{lat},0")
                        continue

                    pop_header_bg = "#d9534f" if r.get('PRIORIDADE') == "Sim" else "#0D256C"
                    pop_prio_txt = "🚨 OBRA PRIORITÁRIA" if r.get('PRIORIDADE') == "Sim" else "📍 Atendimento Padrão"
                    dist_prox = r.get('DISTANCIA_PROXIMO_PONTO_KM', 0.0)
                    
                    extra_rows_list = []
                    for c in cols_exibir:
                        if c.upper() != 'PROTOCOLO':
                            val_html = formata_campo_html(r.get(c, ''))
                            extra_rows_list.append(f"<tr><td style='padding:4px 8px; font-weight:bold; color:#444; border-bottom:1px solid #eee; vertical-align:top; width:35%;'>{html.escape(str(c))}:</td><td style='padding:4px 8px; color:#222; border-bottom:1px solid #eee;'>{val_html}</td></tr>")
                    extra_rows = "".join(extra_rows_list)

                    prot_html = formata_campo_html(r.get('PROTOCOLO', 'N/A'))

                    popup_html = f"""
                    <div style="font-family: Arial, sans-serif; width:320px; border-radius:6px; overflow:hidden; border:1px solid #ccc; background:#fff;">
                        <div style="background:{pop_header_bg}; color:white; padding:10px; font-size:14px; font-weight:bold; text-align:center;">{pop_prio_txt}</div>
                        <div style="padding:10px; background:#fafafa; font-size:13px;">
                            <table style="width:100%; border-collapse:collapse; text-align:left;">
                                <tr><td style="padding:4px 8px; font-weight:bold; color:#444; border-bottom:1px solid #eee; vertical-align:top; width:35%;">Protocolo:</td><td style="padding:4px 8px; color:#222; border-bottom:1px solid #eee;">{prot_html}</td></tr>
                                <tr><td style="padding:4px 8px; font-weight:bold; color:#444; border-bottom:1px solid #eee;">Ordem:</td><td style="padding:4px 8px; color:#222; border-bottom:1px solid #eee;">{r.get('ORDEM', 0)} (Dia {r.get('DIA', 0)})</td></tr>
                                <tr><td style="padding:4px 8px; font-weight:bold; color:#444; border-bottom:1px solid #eee;">Horário:</td><td style="padding:4px 8px; color:#222; border-bottom:1px solid #eee;">{r.get('HORA_INICIO', '')} às {r.get('HORA_FIM', '')}</td></tr>
                                <tr><td style="padding:4px 8px; font-weight:bold; color:#444; border-bottom:1px solid #eee;">Distância Ant.:</td><td style="padding:4px 8px; color:#222; border-bottom:1px solid #eee;">{r.get('DISTANCIA_PONTO_ANTERIOR_KM', 0)} KM</td></tr>
                                <tr><td style="padding:4px 8px; font-weight:bold; color:#444; border-bottom:1px solid #eee;">Distância Próx.:</td><td style="padding:4px 8px; color:#222; border-bottom:1px solid #eee;">{dist_prox} KM</td></tr>
                                {extra_rows}
                            </table>
                        </div>
                    </div>"""
                    
                    is_super = str(r.get('SUPER_PONTO', '')).startswith('SIM')
                    tag_prio = "[PRIORIDADE] " if r.get('PRIORIDADE') == "Sim" else ""
                    
                    if is_super:
                        qtd_str = str(r.get('SUPER_PONTO')).replace('SIM', '').strip()
                        nome_ponto = f"{tag_prio}[{r.get('ORDEM', 0)}] 🏢 SUPER PONTO {qtd_str}"
                        style_url = "#icon-yellow"
                    else:
                        prot_str = str(r.get('PROTOCOLO', 'N/A'))
                        nome_obra = str(r.get('NOME', ''))
                        if nome_obra.lower() == 'nan': nome_obra = ''
                        separador = " - " if nome_obra else ""
                        nome_ponto = f"{tag_prio}[{r.get('ORDEM', 0)}] Prot: {html.escape(prot_str)}{separador}{html.escape(nome_obra)}"
                        style_url = "#icon-red" if r.get('PRIORIDADE') == "Sim" else "#icon-blue"

                    kml_lines.append(f'        <Placemark><name>{nome_ponto}</name><description><![CDATA[{popup_html}]]></description><styleUrl>{style_url}</styleUrl><Point><coordinates>{lon},{lat},0</coordinates></Point></Placemark>')
                    
                    geometria = r.get('ROTA_GEOMETRIA')
                    if isinstance(geometria, list):
                        coords_linha_kml.extend([f"          {pt_lon},{pt_lat},0" for pt_lon, pt_lat in geometria])
                    else:
                        coords_linha_kml.append(f"          {lon},{lat},0")

                kml_str_coords = "\n".join(coords_linha_kml)
                kml_lines.append(f'        <Placemark><name>Traçado Rota</name><styleUrl>#rota-centro-{nome_limpo_base}</styleUrl><LineString><tessellate>1</tessellate><coordinates>\n{kml_str_coords}\n            </coordinates></LineString></Placemark>\n      </Folder>')
            kml_lines.append('    </Folder>')
        kml_lines.append('  </Folder>')
    kml_lines.append('</Document>\n</kml>')
    return "\n".join(kml_lines)

# ==========================================
# 6. TELA PRINCIPAL (UI STREAMLIT)
# ==========================================
def view_roteirizador():
    if "roteamento_concluido" not in st.session_state: st.session_state.roteamento_concluido = False
    if "vrp_status" not in st.session_state: st.session_state.vrp_status = "IDLE"
    if "vrp_state" not in st.session_state: st.session_state.vrp_state = {}
    if "df_routed" not in st.session_state: st.session_state.df_routed = pd.DataFrame()
    if "bases_records" not in st.session_state: st.session_state.bases_records = []
    if "colunas_exibir" not in st.session_state: st.session_state.colunas_exibir = []
    if "col_prioridade" not in st.session_state: st.session_state.col_prioridade = "TIPO NOTA"
    if "colunas_originais" not in st.session_state: st.session_state.colunas_originais = []
    if "config_financeira" not in st.session_state: st.session_state.config_financeira = {}

    status_exec = st.session_state.vrp_status
    is_done = st.session_state.roteamento_concluido

    st.markdown("<h1 class='brand-title'>Plataforma Roteirizadora NIP v1.1</h1>", unsafe_allow_html=True)

    # ---------------------------------------------------------
    # UI DE NAVEGAÇÃO E SIDEBAR (SEMPRE VISÍVEL)
    # ---------------------------------------------------------
    s1_class = "step-item done" if (status_exec != "IDLE" or is_done) else "step-item active"
    s2_class = "step-item done" if (status_exec != "IDLE" or is_done) else "step-item active"
    s3_class = "step-item active" if status_exec in ["RUNNING", "PACKAGING"] else ("step-item done" if is_done else "step-item")
    s4_class = "step-item active" if is_done else "step-item"
    
    st.markdown(f"""
    <div class="stepper-container">
        <div class="{s1_class}">📁 1. Dados e Profiling</div>
        <div class="{s2_class}">⚙️ 2. Filtros Dinâmicos</div>
        <div class="{s3_class}">🚀 3. IA VRP OR-Tools</div>
        <div class="{s4_class}">🎯 4. Resultados e Custos</div>
    </div>
    """, unsafe_allow_html=True)

    is_locked = status_exec != "IDLE" or is_done
    
    with st.sidebar:
        if os.path.exists(LOGO_PATH):
            with open(LOGO_PATH, "rb") as f:
                encoded_logo = base64.b64encode(f.read()).decode()
            st.markdown(
                f'<div style="text-align: center; margin-bottom: 25px;">'
                f'<img src="data:image/png;base64,{encoded_logo}" style="width: 70%; max-width: 180px; pointer-events: none;">'
                f'</div>',
                unsafe_allow_html=True
            )
            
        with st.expander("⚙️ Esforço e Limites Diários", expanded=True):
            tipo_periodo = st.radio("Agrupamento de percurso:", ["Dia", "Semana"], horizontal=True, disabled=is_locked)
            if tipo_periodo == "Semana":
                st.caption("ℹ️ Na visão por Semana, o sistema agrupará 6 dias úteis (Seg-Sáb) em cada aba.")
                
            modo_limite = st.radio("Critério limitador:", ["Quantidade Fixa de Obras", "Carga Horária (Tempo Real)"], disabled=is_locked)
            limite_km_diario = st.slider("Limite de KM por Dia", 0, 500, 500, 5, disabled=is_locked)
            
            if modo_limite == "Quantidade Fixa de Obras":
                obras_por_dia = st.number_input("Obras Previstas por Dia", min_value=1, value=4, step=1, disabled=is_locked)
                limite_periodos = st.number_input(f"Limite total de {tipo_periodo}s", min_value=1, value=5, step=1, disabled=is_locked)
                tempo_medio_obra = 1.5
                velocidade_media_kmh = 30.0
                horas_por_dia = 8.0
            else:
                horas_por_dia = st.number_input("Horas por Dia", min_value=1.0, value=8.0, step=0.5, disabled=is_locked)
                tempo_medio_obra = st.number_input("Tempo de execução/obra (Horas)", min_value=0.1, value=1.5, step=0.1, disabled=is_locked)
                velocidade_media_kmh = st.number_input("Velocidade (km/h)", min_value=10.0, value=30.0, step=5.0, disabled=is_locked)
                limite_periodos = st.number_input(f"Limite total de {tipo_periodo}s", min_value=1, value=5, step=1, disabled=is_locked)
                obras_por_dia = int(horas_por_dia / tempo_medio_obra)
                if obras_por_dia < 1: obras_por_dia = 1

        with st.expander("💰 Custos e Gestão Financeira", expanded=False):
            custo_combustivel = st.number_input("Custo Combustível (R$/L)", min_value=0.0, value=0.0, step=0.1, disabled=is_locked)
            consumo_veiculo = st.number_input("Consumo Frota (Km/L)", min_value=0.0, value=0.0, step=0.5, disabled=is_locked)
            custo_hora_equipe = st.number_input("Hora-Homem da Equipe (R$)", min_value=0.0, value=0.0, step=1.0, disabled=is_locked)
            
        with st.expander("📡 Conexão de Rede (Avançado)", expanded=False):
            url_osrm_base = st.text_input("Endpoint OSRM ⚠️ (NÃO APAGUE OU EDITE):", value="http://router.project-osrm.org", disabled=is_locked)
            st.caption("Este link conecta o sistema à malha viária real de ruas do mundo.")
            
        st.markdown("---")
        timer_placeholder = st.empty()
        
        st.markdown("### 📥 Ações e Arquivos")
        data_atual = datetime.now().strftime("%d_%m_%Y")
        bytes_zip_xl = st.session_state.get('bytes_zip_xl', b"")
        bytes_zip_kml = st.session_state.get('bytes_zip_kml', b"")
        
        botoes_desabilitados = not is_done or st.session_state.df_routed.empty
        
        st.download_button("🌐 1. Baixar Planilhas (ZIP)", data=bytes_zip_xl if bytes_zip_xl else b"vazio", file_name=f"Dados_Estruturados_Roteiro_{data_atual}.zip", mime="application/zip", use_container_width=True, disabled=botoes_desabilitados)
        st.download_button("🗺️ 2. Baixar Mapas (KML)", data=bytes_zip_kml if bytes_zip_kml else b"vazio", file_name=f"Mapas_KML_{data_atual}.zip", mime="application/zip", use_container_width=True, disabled=botoes_desabilitados)
        
        if st.button("🧹 Nova Roteirização", type="primary", use_container_width=True, disabled=botoes_desabilitados): 
            limpar_roteirizador()

    # ---------------------------------------------------------
    # ESTADO 4: RESULTADOS FINAIS (TELA DE SUCESSO)
    # ---------------------------------------------------------
    if is_done and not st.session_state.df_routed.empty:
        st.markdown("## 🎯 Resultados da Otimização")

        st.session_state.df_routed['DISTANCIA_PROXIMO_PONTO_KM'] = st.session_state.df_routed.groupby(['BASE_ATRIBUIDA', 'PERIODO'])['DISTANCIA_PONTO_ANTERIOR_KM'].shift(-1).fillna(0.0)

        df_routed = st.session_state.df_routed.copy()
        bases_records = st.session_state.bases_records
        colunas_exibir = st.session_state.colunas_exibir
        df_real_tasks = df_routed[~df_routed['PROTOCOLO'].isin(['RETORNO_BASE', 'PAUSA_ALMOCO'])]
        
        tot_paradas = len(df_real_tasks)
        tot_obras_reais = sum(len(r['_ORIGINAL_ROWS']) if isinstance(r.get('_ORIGINAL_ROWS'), list) else 1 for _, r in df_real_tasks.iterrows())
        
        tot_equipes = df_routed['BASE_ATRIBUIDA'].nunique()
        tot_km = f"{df_routed['DISTANCIA_PONTO_ANTERIOR_KM'].sum():.1f} km"
        tot_prio = len(df_real_tasks[df_real_tasks['PRIORIDADE'] == 'Sim']) if 'PRIORIDADE' in df_real_tasks else 0

        c_m1, c_m2, c_m3, c_m4 = st.columns(4)
        c_m1.markdown(f'<div class="metric-card" style="border-left: 5px solid #0D256C;"><div class="metric-icon" style="background: rgba(13, 37, 108, 0.12);">📌</div><div class="metric-content"><div class="metric-title">Obras Reais Processadas</div><div class="metric-value">{tot_obras_reais} <span style="font-size:12px;color:#888;">(Em {tot_paradas} Paradas)</span></div></div></div>', unsafe_allow_html=True)
        c_m2.markdown(f'<div class="metric-card" style="border-left: 5px solid #8b5cf6;"><div class="metric-icon" style="background: rgba(139, 92, 246, 0.15);">👥</div><div class="metric-content"><div class="metric-title">Equipes em Campo</div><div class="metric-value">{tot_equipes}</div></div></div>', unsafe_allow_html=True)
        c_m3.markdown(f'<div class="metric-card" style="border-left: 5px solid #55B929;"><div class="metric-icon" style="background: rgba(85, 185, 41, 0.15);">🛣️</div><div class="metric-content"><div class="metric-title">KM Total Projetado</div><div class="metric-value">{tot_km}</div></div></div>', unsafe_allow_html=True)
        c_m4.markdown(f'<div class="metric-card" style="border-left: 5px solid #ef4444;"><div class="metric-icon" style="background: rgba(239, 68, 68, 0.15);">🚨</div><div class="metric-content"><div class="metric-title">Grupos Prioritários</div><div class="metric-value">{tot_prio}</div></div></div>', unsafe_allow_html=True)

        # --- INÍCIO DA AUDITORIA DE META EXATA ---
        cfg_atual = st.session_state.vrp_state.get('config', {})
        
        modo_atual = cfg_atual.get('modo_limite', 'Quantidade Fixa de Obras')
        if modo_atual == "Quantidade Fixa de Obras":
            obras_dia_meta = cfg_atual.get('obras_por_dia', 4)
        else:
            h_dia = cfg_atual.get('horas_por_dia', 8.0)
            t_obra = cfg_atual.get('tempo_medio_obra', 1.5)
            obras_dia_meta = max(1, int(h_dia / t_obra))
            
        dias_multiplier = 6 if cfg_atual.get('tipo_periodo', 'Dia') == 'Semana' else 1
        limite_periodos_meta = cfg_atual.get('limite_periodos', 1)
        
        meta_exata_por_equipe = obras_dia_meta * dias_multiplier * limite_periodos_meta
        tot_equipes_cadastradas = len(st.session_state.bases_records)
        meta_global_exata = meta_exata_por_equipe * tot_equipes_cadastradas
        
        obras_por_equipe = {b['LEVANTADOR']: 0 for b in st.session_state.bases_records}
            
        for _, r in df_real_tasks.iterrows():
            b_name = r['BASE_ATRIBUIDA']
            qtd = len(r.get('_ORIGINAL_ROWS', [1])) if isinstance(r.get('_ORIGINAL_ROWS'), list) else 1
            if b_name in obras_por_equipe:
                obras_por_equipe[b_name] += qtd
                
        equipes_abaixo_meta = {k: v for k, v in obras_por_equipe.items() if v < meta_exata_por_equipe}
        
        if equipes_abaixo_meta:
            motivos = []
            if tot_obras_reais < meta_global_exata:
                motivos.append(f"<b>Falta de Demanda ou Gargalo Geográfico:</b> A meta global exigia {meta_global_exata} obras ({meta_exata_por_equipe} para cada um dos {tot_equipes_cadastradas} levantadores). No entanto, {st.session_state.get('tot_obras_nao_alocadas', 0)} obras ficaram de fora (excesso de demanda total ou localizadas em cidades sem cobertura da base configurada).")
            
            motivos.append("<b>Corte por Tempo/Distância (Segurança VRP):</b> O motor de Inteligência Artificial precisou abortar algumas viagens antes de bater a meta diária da equipe para evitar que o colaborador ultrapassasse o Limite de KM diário configurado, ou que a viagem estourasse o horário limite (após as 18h).")
                
            detalhes_html = "".join([f"<li style='margin-bottom: 4px;'><b>{eq}</b>: Entregou {qtd} de {meta_exata_por_equipe} obras esperadas.</li>" for eq, qtd in equipes_abaixo_meta.items()])
            
            st.markdown(f'''
            <div style="background-color: #fff3cd; color: #856404; padding: 15px; border-left: 5px solid #ffeeba; margin-bottom: 20px; border-radius: 4px;">
                <h4 style="margin-top: 0; color: #856404;">⚠️ Meta Exata Não Atingida ({tot_obras_reais} realizadas de {meta_global_exata} esperadas)</h4>
                <p style="margin-bottom: 10px;">O sistema tem a meta de alocar a quantidade exata para cada levantador, mas interveio ativando a proteção de jornada pelos seguintes motivos técnicos:</p>
                <ul style="margin-bottom: 15px;">
                    {"".join([f"<li style='margin-bottom: 5px;'>{m}</li>" for m in motivos])}
                </ul>
                <details>
                    <summary style="cursor: pointer; font-weight: bold; padding: 5px; background: rgba(0,0,0,0.05); border-radius: 4px;">Ver Levantadores que não atingiram a meta ({len(equipes_abaixo_meta)} equipes)</summary>
                    <ul style="margin-top: 10px;">
                        {detalhes_html}
                    </ul>
                </details>
            </div>
            ''', unsafe_allow_html=True)
        else:
            st.success(f"✅ **Alocação Perfeita:** Todas as equipes atingiram a meta exata de {meta_exata_por_equipe} obras cada, totalizando {meta_global_exata} serviços roteirizados!")
        # --- FIM DA AUDITORIA DE META EXATA ---

        cf_comb = st.session_state.config_financeira.get('custo_combustivel', 0.0)
        cf_cons = st.session_state.config_financeira.get('consumo_veiculo', 0.0)
        cf_hora = st.session_state.config_financeira.get('custo_hora_equipe', 0.0)
        
        mostrar_financeiro = (cf_comb > 0) or (cf_hora > 0)

        if mostrar_financeiro:
            tot_km_val = df_routed['DISTANCIA_PONTO_ANTERIOR_KM'].sum()
            litros_gastos = tot_km_val / cf_cons if cf_cons > 0 else 0
            custo_total_combustivel = litros_gastos * cf_comb
            
            df_financeiro = df_routed.copy()
            df_financeiro['_HORA_INICIO_DT'] = pd.to_datetime(df_financeiro['_HORA_INICIO_DT'])
            df_financeiro['_HORA_FIM_DT'] = pd.to_datetime(df_financeiro['_HORA_FIM_DT'])
            
            custo_total_mao_de_obra = 0.0
            horas_totais = 0.0
            
            for (eq, periodo_f), group in df_financeiro.groupby(['BASE_ATRIBUIDA', 'PERIODO']):
                h_inicio = group['_HORA_INICIO_DT'].min()
                h_fim = group['_HORA_FIM_DT'].max()
                h_trab = (h_fim - h_inicio).total_seconds() / 3600.0
                horas_totais += h_trab
                custo_total_mao_de_obra += (h_trab * cf_hora)
                
            custo_operacao_total = custo_total_combustivel + custo_total_mao_de_obra
            custo_por_obra = custo_operacao_total / tot_obras_reais if tot_obras_reais > 0 else 0.0

            c_fin1, c_fin2, c_fin3, c_fin4 = st.columns(4)
            c_fin1.markdown(f'<div class="metric-card" style="border-left: 5px solid #f59e0b;"><div class="metric-icon" style="background: rgba(245, 158, 11, 0.15);">⛽</div><div class="metric-content"><div class="metric-title">Combustível Estimado</div><div class="metric-value">R$ {formatar_moeda(custo_total_combustivel)}</div></div></div>', unsafe_allow_html=True)
            c_fin2.markdown(f'<div class="metric-card" style="border-left: 5px solid #8b5cf6;"><div class="metric-icon" style="background: rgba(139, 92, 246, 0.15);">👷</div><div class="metric-content"><div class="metric-title">Mão de Obra ({horas_totais:.1f}h)</div><div class="metric-value">R$ {formatar_moeda(custo_total_mao_de_obra)}</div></div></div>', unsafe_allow_html=True)
            c_fin3.markdown(f'<div class="metric-card" style="border-left: 5px solid #ef4444;"><div class="metric-icon" style="background: rgba(239, 68, 68, 0.15);">💲</div><div class="metric-content"><div class="metric-title">Custo Total Operação</div><div class="metric-value">R$ {formatar_moeda(custo_operacao_total)}</div></div></div>', unsafe_allow_html=True)
            c_fin4.markdown(f'<div class="metric-card" style="border-left: 5px solid #55B929;"><div class="metric-icon" style="background: rgba(85, 185, 41, 0.15);">📊</div><div class="metric-content"><div class="metric-title">Custo Médio por Obra</div><div class="metric-value">R$ {formatar_moeda(custo_por_obra)}</div></div></div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown("### 🗺️ Mapa Geográfico de Rotas")
        mapa = folium.Map(location=[df_routed['LATITUDE'].mean(), df_routed['LONGITUDE'].mean()], zoom_start=8) if not df_routed.empty else folium.Map(location=[-5.2, -45.0], zoom_start=7)
        
        cores_folium = ['#e6194b', '#00bcd4', '#3f51b5', '#009688', '#ff9800', '#9c27b0', '#cddc39', '#e91e63', '#ffeb3b', '#795548']
        lista_bases_mapa = df_routed['BASE_ATRIBUIDA'].unique().tolist()
        
        heat_data = [[r['LATITUDE'], r['LONGITUDE']] for _, r in df_real_tasks.iterrows()]
        HeatMap(heat_data, name="🔥 Mapa de Calor (Demandas)", radius=15, blur=10).add_to(mapa)
        marker_cluster = MarkerCluster(name="Obras (Agrupadas)").add_to(mapa)
        
        for base_nome in lista_bases_mapa:
            idx_cor = lista_bases_mapa.index(base_nome)
            cor_rota = cores_folium[idx_cor % len(cores_folium)]
            df_base_rota = df_routed[df_routed['BASE_ATRIBUIDA'] == base_nome]
            base_ref = next((b for b in bases_records if b['LEVANTADOR'] == base_nome), None)
            b_lat, b_lon = float(str(base_ref['LATITUDE']).replace(',','.')), float(str(base_ref['LONGITUDE']).replace(',','.'))
            folium.Marker([b_lat, b_lon], icon=folium.Icon(color='black', icon='home', prefix='fa'), tooltip=f"Base: {base_nome}").add_to(mapa)
            
            fg_linhas = folium.FeatureGroup(name=f"Rota: {base_nome}", show=False)
            
            for periodo_val in df_base_rota['PERIODO'].unique():
                df_periodo = df_base_rota[df_base_rota['PERIODO'] == periodo_val]
                
                pontos_linha_folium = []
                for _, r in df_periodo.iterrows():
                    if isinstance(r.get('ROTA_GEOMETRIA'), list):
                        for lon, lat in r['ROTA_GEOMETRIA']: pontos_linha_folium.append([lat, lon]) 
                            
                folium.PolyLine(pontos_linha_folium, color='black', weight=7, opacity=0.9).add_to(fg_linhas)
                folium.PolyLine(pontos_linha_folium, color=cor_rota, weight=3, opacity=1.0).add_to(fg_linhas)
                
                for r in df_periodo.to_dict('records'):
                    if r.get('PROTOCOLO') in ['RETORNO_BASE', 'PAUSA_ALMOCO']: continue
                    icone = identificar_icone_folium(r, df_routed.columns)
                    
                    is_super = str(r.get('SUPER_PONTO', '')).startswith('SIM')
                    if is_super:
                        cor_icone = 'orange'
                    else:
                        cor_icone = 'red' if r.get('PRIORIDADE') == "Sim" else 'blue'
                    
                    pop_header_bg = "#d9534f" if r.get('PRIORIDADE') == "Sim" else "#0D256C"
                    pop_prio_txt = "🚨 OBRA PRIORITÁRIA" if r.get('PRIORIDADE') == "Sim" else "📍 Atendimento Padrão"
                    
                    dist_prox = r.get('DISTANCIA_PROXIMO_PONTO_KM', 0.0)
                    
                    extra_rows_list = []
                    for c in colunas_exibir:
                        if c.upper() != 'PROTOCOLO':
                            val_html = formata_campo_html(r.get(c, ''))
                            extra_rows_list.append(f"<tr><td style='padding:3px 6px; font-weight:bold; color:#555; vertical-align:top; width:35%;'>{html.escape(str(c))}:</td><td style='padding:3px 6px; color:#333;'>{val_html}</td></tr>")
                    extra_rows = "".join(extra_rows_list)

                    prot_html = formata_campo_html(r.get('PROTOCOLO', 'N/A'))

                    popup_html = f"""
                    <div style="font-family:sans-serif; width:280px; border-radius:8px; overflow:hidden; box-shadow:0 2px 5px rgba(0,0,0,0.15);">
                        <div style="background:{pop_header_bg}; color:white; padding:8px 10px; font-size:13px; font-weight:bold;">{pop_prio_txt}</div>
                        <div style="padding:10px; background:#fafafa; font-size:12px;">
                            <table style="width:100%; border-collapse:collapse;">
                                <tr><td style="padding:3px 6px; font-weight:bold; color:#555; vertical-align:top; width:35%;">Protocolo:</td><td style="padding:3px 6px; color:#333;">{prot_html}</td></tr>
                                <tr><td style="padding:3px 6px; font-weight:bold; color:#555;">Ordem:</td><td style="padding:3px 6px; color:#333;">{r.get('ORDEM', 0)} ({tipo_periodo} {r.get('PERIODO', 0)})</td></tr>
                                <tr><td style="padding:3px 6px; font-weight:bold; color:#555;">Horário:</td><td style="padding:3px 6px; color:#333;">{r.get('HORA_INICIO', '')} às {r.get('HORA_FIM', '')}</td></tr>
                                <tr><td style="padding:3px 6px; font-weight:bold; color:#555;">Distância Ant.:</td><td style="padding:3px 6px; color:#333;">{r.get('DISTANCIA_PONTO_ANTERIOR_KM', 0)} KM</td></tr>
                                <tr><td style="padding:3px 6px; font-weight:bold; color:#555;">Distância Próx.:</td><td style="padding:3px 6px; color:#333;">{dist_prox} KM</td></tr>
                                {extra_rows}
                            </table>
                        </div>
                    </div>"""
                    folium.Marker([r['LATITUDE'], r['LONGITUDE']], icon=folium.Icon(color=cor_icone, icon=icone), popup=folium.Popup(popup_html, max_width=300)).add_to(marker_cluster)
            
            fg_linhas.add_to(mapa)
            
        folium.LayerControl().add_to(mapa)
        st_folium(mapa, use_container_width=True, height=550, returned_objects=[])

        st.markdown("<br>", unsafe_allow_html=True)
        tab_dados, tab_gantt = st.tabs(["📊 Dados Tabulares e Arquivos", "⏱️ Timeline Interativa (Gantt)"])

        with tab_dados:
            st.markdown("#### Detalhamento de Rotas")
            df_display = st.session_state.df_routed.drop(columns=['ROTA_GEOMETRIA', '_HORA_INICIO_DT', '_HORA_FIM_DT', '_ORIGINAL_ROWS'], errors='ignore')
            
            df_editado_ui = st.data_editor(
                df_display, use_container_width=True, height=400,
                column_config={ 
                    "LATITUDE": st.column_config.NumberColumn(disabled=True), "LONGITUDE": st.column_config.NumberColumn(disabled=True),
                    "DISTANCIA_PONTO_ANTERIOR_KM": st.column_config.ProgressColumn("Dist. Anterior (KM)", format="%.2f", min_value=0, max_value=30), 
                    "TEMPO_VIAGEM_MINUTOS": st.column_config.ProgressColumn("Tempo de Viagem (Min)", format="%.1f", min_value=0, max_value=60)
                }
            )

        with tab_gantt:
            st.markdown("#### Linha do Tempo das Equipes")
            df_gantt = df_routed.copy()
            
            def label_tarefa(row):
                if row['PROTOCOLO'] == 'RETORNO_BASE': return 'Deslocamento/Retorno'
                if row['PROTOCOLO'] == 'PAUSA_ALMOCO': return 'Pausa para Almoço'
                prio = "🚨 " if row.get('PRIORIDADE') == 'Sim' else ""
                return f"{prio}Obra Operacional"
                
            df_gantt['TAREFA'] = df_gantt.apply(label_tarefa, axis=1)
            
            df_gantt['START_PLOT'] = pd.to_datetime(df_gantt['_HORA_INICIO_DT'])
            df_gantt['END_PLOT'] = pd.to_datetime(df_gantt['_HORA_FIM_DT'])
            
            mask_retorno = df_gantt['PROTOCOLO'] == 'RETORNO_BASE'
            df_gantt.loc[mask_retorno, 'END_PLOT'] = df_gantt.loc[mask_retorno, 'END_PLOT'] + pd.Timedelta(minutes=5)
            
            periodos_disp = sorted(df_gantt['PERIODO'].unique())
            if periodos_disp:
                p_sel = st.selectbox("Selecione o Período:", periodos_disp)
                df_plot = df_gantt[df_gantt['PERIODO'] == p_sel]
                
                fig = px.timeline(
                    df_plot, 
                    x_start="START_PLOT", x_end="END_PLOT", y="BASE_ATRIBUIDA", color="TAREFA",
                    hover_name="PROTOCOLO",
                    color_discrete_map={'Deslocamento/Retorno': '#6c757d', 'Pausa para Almoço': '#f39c12', 'Obra Operacional': '#0D256C', '🚨 Obra Operacional': '#e74c3c'},
                    height=300 + (len(df_plot['BASE_ATRIBUIDA'].unique()) * 30)
                )
                fig.update_yaxes(autorange="reversed", title="")
                fig.layout.xaxis.tickformat = '%H:%M'
                fig.update_layout(
                    xaxis_title="Horário do Dia", 
                    margin=dict(l=0, r=0, t=30, b=0)
                )
                st.plotly_chart(fig, use_container_width=True)

        return 

    # ---------------------------------------------------------
    # ESTADO 3.1: MOTOR IA (VRP) E BALANCEAMENTO DE CARGA
    # ---------------------------------------------------------
    def fetch_geom_wrapper(item):
        try:
            geom, dur_sec = obter_rota_ruas(item['lat_ant'], item['lon_ant'], item['lat_atual'], item['lon_atual'], url_osrm_base, cfg['velocidade_media_kmh'])
            return geom, dur_sec
        except Exception:
            coords = np.array([[item['lat_ant'], item['lon_ant']], [item['lat_atual'], item['lon_atual']]])
            dist_m = calcular_matriz_distancias_numpy(coords)[0][1]
            return [[item['lon_ant'], item['lat_ant']], [item['lon_atual'], item['lat_atual']]], (dist_m / 1000.0 / cfg['velocidade_media_kmh']) * 3600

    if status_exec in ["RUNNING"]:
        st.markdown("## 🚀 Execução do Motor de Inteligência (OR-Tools VRP)")
        st.markdown("Calculando Matrizes Vetoriais e Otimizando Rotas...")
        
        if st.button("⏹️ Abortar Execução", use_container_width=True): limpar_roteirizador()
            
        state = st.session_state.vrp_state
        cfg = state['config']
        b_names = state['b_names']
        total_equipes = len(b_names)
        
        progress_bar = st.progress(0.0)
        status_text = st.empty()
        timer_placeholder = st.empty()
        
        try:
            tempo_processamento = 0.0
            routed_data_final = []
            df_todas_bases_ativas = pd.DataFrame(st.session_state.bases_records)
            unvisited = state['unvisited']
            
            for b_idx, b_name in enumerate(b_names):
                start_iter = time.time()
                progresso = b_idx / total_equipes if total_equipes > 0 else 1.0
                progress_bar.progress(progresso)
                status_text.info(f"🧠 IA Analisando nós e traçando rotas para **{b_name}**... ({b_idx + 1}/{total_equipes})")
                
                with timer_placeholder.container():
                    if b_idx > 0:
                        avg = tempo_processamento / b_idx
                        restantes = total_equipes - b_idx
                        est_rem = avg * restantes
                        m, s = divmod(int(est_rem), 60)
                        h, m = divmod(m, 60)
                        time_str = f"{h:02d}h {m:02d}m {s:02d}s" if h > 0 else f"{m:02d}m {s:02d}s"
                        
                        st.markdown("### ⏱️ Tempo Restante Estimado")
                        st.markdown(f"""
                        <div style="padding: 0.75rem 1rem; border-radius: 0.5rem; background-color: rgba(85, 185, 41, 0.15); color: #2e7d32; border: 1px solid rgba(85, 185, 41, 0.3); display: flex; align-items: center;">
                            <span style="font-size:1.5rem; margin-right:12px;">⏳</span> 
                            <strong style="font-size:1.2rem;">{time_str}</strong>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.markdown("### ⏱️ Tempo Restante Estimado")
                        st.markdown(f"""
                        <div style="padding: 0.75rem 1rem; border-radius: 0.5rem; background-color: rgba(13, 37, 108, 0.12); color: #0D256C; border: 1px solid rgba(13, 37, 108, 0.25); display: flex; align-items: center;">
                            <span style="font-size:1.2rem; margin-right:10px;">🔄</span> 
                            <span>Calculando estimativa...</span>
                        </div>
                        """, unsafe_allow_html=True)

                base_ref = df_todas_bases_ativas[df_todas_bases_ativas['LEVANTADOR'] == b_name].iloc[0]
                if pd.isna(base_ref.get('LATITUDE')):
                    continue
                    
                base_lat, base_lon = float(base_ref['LATITUDE']), float(base_ref['LONGITUDE'])
                obras_equipe = unvisited[unvisited['BASE_ATRIBUIDA'] == b_name].to_dict('records')
                
                if obras_equipe:
                    coords_dict = {}
                    for o in obras_equipe:
                        k = (round(float(o['LATITUDE']), 4), round(float(o['LONGITUDE']), 4))
                        if k not in coords_dict:
                            coords_dict[k] = []
                        coords_dict[k].append(o)
                        
                    macro_obras = []
                    for k, lista in coords_dict.items():
                        tem_prio = any(x.get('PRIORIDADE') == 'Sim' for x in lista)
                        rep = lista[0].copy()
                        rep['PRIORIDADE'] = 'Sim' if tem_prio else 'Não'
                        rep['_sub_obras'] = lista
                        macro_obras.append(rep)

                    prio_macros = [m for m in macro_obras if m['PRIORIDADE'] == 'Sim']
                    comum_macros = [m for m in macro_obras if m['PRIORIDADE'] != 'Sim']
                    
                    ordered_macros = []
                    if prio_macros:
                        ordered_macros.extend(resolver_tsp_ortools(prio_macros, base_lat, base_lon, url_osrm_base))
                    if comum_macros:
                        ordered_macros.extend(resolver_tsp_ortools(comum_macros, base_lat, base_lon, url_osrm_base))
                        
                    ordered_tasks = []
                    for macro in ordered_macros:
                        subs = sorted(macro['_sub_obras'], key=lambda x: 0 if x.get('PRIORIDADE') == 'Sim' else 1)
                        ordered_tasks.extend(subs)
                    
                    rotas_flat = []
                    dia_absoluto = 1
                    semana_atual = 1
                    dia_da_semana = 1
                    obras_no_periodo_macro = 0
                    
                    agora_dt = datetime.now()
                    data_base_inicio = agora_dt.replace(hour=8, minute=0, second=0, microsecond=0)
                    data_base_almoco = agora_dt.replace(hour=12, minute=0, second=0, microsecond=0)
                    
                    def iniciar_dia(dia_abs):
                        return {
                            'lat': base_lat, 'lon': base_lon,
                            'time': data_base_inicio + pd.Timedelta(days=dia_abs - 1),
                            'obras_hoje': 0,
                            'km_hoje': 0.0,
                            'lunch': False
                        }
                        
                    estado = iniciar_dia(dia_absoluto)
                    
                    for obra in ordered_tasks:
                        qtd_real = len(obra.get('_ORIGINAL_ROWS', [1])) if isinstance(obra.get('_ORIGINAL_ROWS'), list) else 1
                        
                        if cfg['tipo_periodo'] == "Dia" and dia_absoluto > cfg['limite_periodos']: break
                        if cfg['tipo_periodo'] == "Semana" and semana_atual > cfg['limite_periodos']: break

                        viagem_km = haversine_vectorized(estado['lat'], estado['lon'], obra['LATITUDE'], obra['LONGITUDE'])
                        
                        if viagem_km < 0.05 and estado['obras_hoje'] > 0:
                            viagem_min = 0.0
                            exec_min = 30.0 
                        else:
                            viagem_min = (viagem_km / cfg['velocidade_media_kmh']) * 60
                            exec_min = cfg['tempo_medio_obra'] * 60
                        
                        chegada_prevista = estado['time'] + pd.Timedelta(minutes=viagem_min)
                        
                        if chegada_prevista.hour >= 12 and not estado['lunch']:
                            lunch_start = max(estado['time'], data_base_almoco + pd.Timedelta(days=dia_absoluto - 1))
                            lunch_end = lunch_start + pd.Timedelta(hours=1)
                            
                            rotas_flat.append({
                                'obra': None, 'is_lunch': True, 'is_retorno': False,
                                'lat_ant': estado['lat'], 'lon_ant': estado['lon'],
                                'lat_atual': estado['lat'], 'lon_atual': estado['lon'],
                                'semana': semana_atual, 'dia': dia_absoluto,
                                'hora_inicio': lunch_start, 'hora_fim': lunch_end,
                                'viagem_min': 0.0, 'dist_km': 0.0
                            })
                            estado['time'] = lunch_end
                            estado['lunch'] = True
                            chegada_prevista = estado['time'] + pd.Timedelta(minutes=viagem_min)
                            
                        fim_previsto = chegada_prevista + pd.Timedelta(minutes=exec_min)
                        
                        virar_dia = False
                        
                        # Correção de Calendário (Bug Invisível Consertado)
                        if fim_previsto.hour >= 18 or fim_previsto.date() > estado['time'].date():
                            virar_dia = True
                            
                        if cfg['modo_limite'] == "Quantidade Fixa de Obras" and obras_no_periodo_macro >= cfg['obras_por_dia']:
                            virar_dia = True
                            
                        if estado.get('km_hoje', 0.0) + viagem_km > cfg.get('limite_km_diario', 500):
                            virar_dia = True
                                
                        if virar_dia:
                            dist_ret = haversine_vectorized(estado['lat'], estado['lon'], base_lat, base_lon)
                            viagem_ret = (dist_ret / cfg['velocidade_media_kmh']) * 60
                            ret_fim = estado['time'] + pd.Timedelta(minutes=viagem_ret)
                            
                            rotas_flat.append({
                                'obra': None, 'is_lunch': False, 'is_retorno': True,
                                'lat_ant': estado['lat'], 'lon_ant': estado['lon'],
                                'lat_atual': base_lat, 'lon_atual': base_lon,
                                'semana': semana_atual, 'dia': dia_absoluto,
                                'hora_inicio': estado['time'], 'hora_fim': ret_fim,
                                'viagem_min': viagem_ret, 'dist_km': dist_ret
                            })
                            
                            dia_absoluto += 1
                            obras_no_periodo_macro = 0
                            
                            if cfg['tipo_periodo'] == "Semana":
                                dia_da_semana += 1
                                if dia_da_semana > 6:
                                    semana_atual += 1
                                    dia_da_semana = 1
                                    
                            if cfg['tipo_periodo'] == "Dia" and dia_absoluto > cfg['limite_periodos']: break
                            if cfg['tipo_periodo'] == "Semana" and semana_atual > cfg['limite_periodos']: break
                                
                            estado = iniciar_dia(dia_absoluto)
                            
                            viagem_km = haversine_vectorized(estado['lat'], estado['lon'], obra['LATITUDE'], obra['LONGITUDE'])
                            if viagem_km < 0.05 and estado['obras_hoje'] > 0:
                                viagem_min = 0.0
                                exec_min = 30.0 
                            else:
                                viagem_min = (viagem_km / cfg['velocidade_media_kmh']) * 60
                                exec_min = cfg['tempo_medio_obra'] * 60
                            
                            chegada_prevista = estado['time'] + pd.Timedelta(minutes=viagem_min)
                            fim_previsto = chegada_prevista + pd.Timedelta(minutes=exec_min)
                            
                        rotas_flat.append({
                            'obra': obra, 'is_lunch': False, 'is_retorno': False,
                            'lat_ant': estado['lat'], 'lon_ant': estado['lon'],
                            'lat_atual': obra['LATITUDE'], 'lon_atual': obra['LONGITUDE'],
                            'semana': semana_atual, 'dia': dia_absoluto,
                            'hora_inicio': chegada_prevista, 'hora_fim': fim_previsto,
                            'viagem_min': viagem_min, 'dist_km': viagem_km
                        })
                        estado['lat'] = obra['LATITUDE']
                        estado['lon'] = obra['LONGITUDE']
                        estado['time'] = fim_previsto
                        estado['obras_hoje'] += qtd_real
                        estado['km_hoje'] += viagem_km
                        obras_no_periodo_macro += qtd_real

                    if estado['obras_hoje'] > 0 and not (cfg['tipo_periodo'] == "Dia" and dia_absoluto > cfg['limite_periodos']) and not (cfg['tipo_periodo'] == "Semana" and semana_atual > cfg['limite_periodos']):
                        dist_ret = haversine_vectorized(estado['lat'], estado['lon'], base_lat, base_lon)
                        viagem_ret = (dist_ret / cfg['velocidade_media_kmh']) * 60
                        ret_fim = estado['time'] + pd.Timedelta(minutes=viagem_ret)
                        rotas_flat.append({
                            'obra': None, 'is_lunch': False, 'is_retorno': True,
                            'lat_ant': estado['lat'], 'lon_ant': estado['lon'],
                            'lat_atual': base_lat, 'lon_atual': base_lon,
                            'semana': semana_atual, 'dia': dia_absoluto,
                            'hora_inicio': estado['time'], 'hora_fim': ret_fim,
                            'viagem_min': viagem_ret, 'dist_km': dist_ret
                        })

                    with ThreadPoolExecutor(max_workers=10) as executor:
                        geoms_and_durs = list(executor.map(fetch_geom_wrapper, rotas_flat))

                    ordem_global = 1
                    for item, (geom, dur_sec) in zip(rotas_flat, geoms_and_durs):
                        periodo_val = item['semana'] if cfg['tipo_periodo'] == "Semana" else item['dia']
                        
                        if item['is_lunch']:
                            routed_data_final.append({
                                'PROTOCOLO': 'PAUSA_ALMOCO', 'NOME': '🍔 ALMOÇO DA EQUIPE', 
                                'LATITUDE': item['lat_atual'], 'LONGITUDE': item['lon_atual'],
                                'BASE_ATRIBUIDA': b_name, 'ORDEM': ordem_global, 
                                'SEMANA': item['semana'],
                                'DIA': item['dia'], 
                                'PERIODO': periodo_val,
                                'DISTANCIA_PONTO_ANTERIOR_KM': 0.0, 'TEMPO_VIAGEM_MINUTOS': 0.0,
                                'ROTA_GEOMETRIA': [[item['lon_atual'], item['lat_atual']], [item['lon_atual'], item['lat_atual']]],
                                'PRIORIDADE': 'Não',
                                'HORA_INICIO': item['hora_inicio'].strftime('%H:%M'),
                                'HORA_FIM': item['hora_fim'].strftime('%H:%M'),
                                '_HORA_INICIO_DT': item['hora_inicio'], '_HORA_FIM_DT': item['hora_fim']
                            })
                        elif item['is_retorno']:
                            routed_data_final.append({
                                'PROTOCOLO': 'RETORNO_BASE', 'NOME': 'BASE_RETORNO', 
                                'LATITUDE': item['lat_atual'], 'LONGITUDE': item['lon_atual'],
                                'BASE_ATRIBUIDA': b_name, 'ORDEM': ordem_global, 
                                'SEMANA': item['semana'],
                                'DIA': item['dia'], 
                                'PERIODO': periodo_val,
                                'DISTANCIA_PONTO_ANTERIOR_KM': round(item['dist_km'], 2), 
                                'TEMPO_VIAGEM_MINUTOS': round(item['viagem_min'], 1),
                                'ROTA_GEOMETRIA': geom, 'PRIORIDADE': 'Não',
                                'HORA_INICIO': item['hora_inicio'].strftime('%H:%M'),
                                'HORA_FIM': item['hora_fim'].strftime('%H:%M'),
                                '_HORA_INICIO_DT': item['hora_inicio'], '_HORA_FIM_DT': item['hora_fim']
                            })
                        else:
                            obra = item['obra']
                            obra['ORDEM'] = ordem_global
                            obra['SEMANA'] = item['semana']
                            obra['DIA'] = item['dia']
                            obra['PERIODO'] = periodo_val
                            obra['DISTANCIA_PONTO_ANTERIOR_KM'] = round(item['dist_km'], 2)
                            obra['TEMPO_VIAGEM_MINUTOS'] = round(item['viagem_min'], 1)
                            obra['ROTA_GEOMETRIA'] = geom
                            obra['HORA_INICIO'] = item['hora_inicio'].strftime('%H:%M')
                            obra['HORA_FIM'] = item['hora_fim'].strftime('%H:%M')
                            obra['_HORA_INICIO_DT'] = item['hora_inicio']
                            obra['_HORA_FIM_DT'] = item['hora_fim']
                            
                            routed_data_final.append(obra)
                        ordem_global += 1

                tempo_processamento += (time.time() - start_iter)
                
            status_text.success("✅ Matrizes Resolvidas! Preparando empacotamento...")
            progress_bar.progress(1.0)
            
            df_final_route = pd.DataFrame(routed_data_final)
            if not df_final_route.empty:
                df_final_route['DISTANCIA_PROXIMO_PONTO_KM'] = df_final_route.groupby(['BASE_ATRIBUIDA', 'PERIODO'])['DISTANCIA_PONTO_ANTERIOR_KM'].shift(-1).fillna(0.0)
                
            st.session_state.df_routed = df_final_route
            st.session_state.vrp_status = "PACKAGING"
            time.sleep(1)
            tentar_rerun()
            
        except Exception as e:
            st.error(f"🚨 ERRO CRÍTICO NO MOTOR DE INTELIGÊNCIA: {e}")
            import traceback
            st.code(traceback.format_exc())
            st.session_state.vrp_status = "IDLE"
            if st.button("⬅️ Voltar e Tentar Novamente"): limpar_roteirizador()
            return

    # ---------------------------------------------------------
    # ESTADO 3.2: EMPACOTAMENTO FINAL (ZIP EXCEL E KML)
    # ---------------------------------------------------------
    if status_exec == "PACKAGING":
        st.markdown("## 📦 Finalizando Otimização e Empacotando Arquivos...")
        progress_bar = st.progress(0.0)
        status_text = st.empty()
        
        try:
            status_text.info("Gerando Planilhas Excel...")
            df_r = st.session_state.df_routed
            col_prio = st.session_state.get('col_prioridade', 'TIPO NOTA')
            cols_orig = st.session_state.get('colunas_originais', [])
            
            excel_bytes = gerar_excel_bytes(df_r, col_prio, cols_orig)
            
            buf_zip_xl = io.BytesIO()
            with zipfile.ZipFile(buf_zip_xl, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.writestr("Roteiros_Detalhados.xlsx", excel_bytes)
                
            st.session_state.bytes_zip_xl = buf_zip_xl.getvalue()
            progress_bar.progress(0.5)
            
            status_text.info("Gerando Mapas de Navegação (KML)...")
            
            buf_zip_kml = io.BytesIO()
            with zipfile.ZipFile(buf_zip_kml, 'w', zipfile.ZIP_DEFLATED) as zf:
                for base_nome in df_r['BASE_ATRIBUIDA'].unique():
                    kml_content = gerar_kml_agrupado(
                        df_r, 
                        st.session_state.bases_records, 
                        f"Rota_{base_nome}", 
                        st.session_state.colunas_exibir, 
                        [base_nome], 
                        st.session_state.tipo_periodo
                    )
                    nome_arq = re.sub(r'[^A-Za-z0-9_]', '', str(base_nome))
                    zf.writestr(f"Rota_{nome_arq}.kml", kml_content.encode('utf-8'))
                    
            st.session_state.bytes_zip_kml = buf_zip_kml.getvalue()
            progress_bar.progress(1.0)
            status_text.success("✅ Tudo pronto!")
            
            st.session_state.roteamento_concluido = True
            st.session_state.vrp_status = "IDLE"
            time.sleep(1)
            tentar_rerun()
            
        except Exception as e:
            st.error(f"🚨 ERRO NO EMPACOTAMENTO: {e}")
            import traceback
            st.code(traceback.format_exc())
            st.session_state.vrp_status = "IDLE"
            if st.button("⬅️ Voltar"): limpar_roteirizador()
            return

    # ---------------------------------------------------------
    # ESTADO 1 E 2: UPLOAD E FILTROS INICIAIS
    # ---------------------------------------------------------
    if status_exec == "IDLE" and not is_done:
        col_up_1, col_up_2 = st.columns(2)

        with col_up_1:
            st.markdown("### 👥 1. Levantadores Principais")
            df_bases = pd.DataFrame()

            with st.container(border=True):
                base_file = st.file_uploader("Suba a planilha Levantadores_MA", type=["xlsx", "xls"])
            
            if base_file:
                try:
                    df_bases_temp_ui = ler_planilha_cached(base_file.getvalue())
                    df_bases_temp_ui.columns = normalize_cols(df_bases_temp_ui.columns)
                    if 'LEVANTADOR' not in df_bases_temp_ui.columns:
                        for p_nome in ['NOME', 'TECNICO', 'EQUIPE', 'COLABORADOR']:
                            if p_nome in df_bases_temp_ui.columns:
                                df_bases_temp_ui = df_bases_temp_ui.rename(columns={p_nome: 'LEVANTADOR'})
                                break
                    if 'LEVANTADOR' in df_bases_temp_ui.columns:
                        opcoes_levs = sorted([str(x) for x in df_bases_temp_ui['LEVANTADOR'].dropna().unique().tolist() if str(x).upper().strip() != 'SEM LEVANTADOR'])
                        
                        levs_selecionados = st.multiselect("Selecione as Equipes Principais:", opcoes_levs, default=opcoes_levs)
                        
                        if levs_selecionados:
                            df_bases = df_bases_temp_ui[df_bases_temp_ui['LEVANTADOR'].isin(levs_selecionados)].copy()
                            
                            if 'LATITUDE' in df_bases.columns and 'LONGITUDE' in df_bases.columns:
                                df_bases['LATITUDE'] = pd.to_numeric(df_bases['LATITUDE'].astype(str).str.replace(',', '.'), errors='coerce')
                                df_bases['LONGITUDE'] = pd.to_numeric(df_bases['LONGITUDE'].astype(str).str.replace(',', '.'), errors='coerce')
                            elif 'RESIDENCIA' in df_bases.columns or 'MUNICIPIO' in df_bases.columns:
                                col_ref = 'RESIDENCIA' if 'RESIDENCIA' in df_bases.columns else 'MUNICIPIO'
                                muns_unicos = df_bases[col_ref].dropna().unique()
                                mapa_coords = {}
                                with st.spinner("🌍 Buscando coordenadas base no satélite..."):
                                    for mun in muns_unicos:
                                        lat, lon = obter_coordenadas_municipio_cached(mun)
                                        mapa_coords[mun] = (lat, lon)
                                df_bases['LATITUDE'] = df_bases[col_ref].map(lambda x: mapa_coords.get(x, (np.nan, np.nan))[0])
                                df_bases['LONGITUDE'] = df_bases[col_ref].map(lambda x: mapa_coords.get(x, (np.nan, np.nan))[1])
                            else:
                                df_bases['LATITUDE'] = np.nan
                                df_bases['LONGITUDE'] = np.nan
                                
                            df_bases = df_bases.dropna(subset=['LATITUDE', 'LONGITUDE'])
                            df_bases['TIPO_EQUIPE'] = 'PRINCIPAL'
                    else:
                        st.error("❌ A planilha não possui a coluna 'LEVANTADOR'.")
                except Exception as e:
                    st.error(f"Erro ao ler a planilha: {e}")

            st.markdown("##### Regra de Atribuição Territorial")
            tipo_atribuicao = st.radio("Regra", ["Por Municípios Atendidos (Lê texto da planilha)", "Por Proximidade Geográfica das Coordenadas (Ignora texto)", "Clusterização Inteligente por IA (K-Means)"], index=0, label_visibility="collapsed")

        with col_up_2:
            st.markdown("### 📁 2. Upload de Demandas (Obras)")
            
            with st.container(border=True):
                task_files = st.file_uploader("1️⃣ Base Principal", type=["xlsx", "xls"], accept_multiple_files=True)
            
            with st.container(border=True):
                status_file = st.file_uploader("2️⃣ Planilha Atualizada SharePoint (Opcional)", type=["xlsx", "xls"])
            
            df_status_upload = pd.DataFrame()
            coluna_status_selecionada = None
            if status_file:
                try:
                    df_status_upload = ler_planilha_cached(status_file.getvalue())
                    cols_status = df_status_upload.columns.tolist()
                    coluna_status_selecionada = st.selectbox("📌 Coluna Status?", cols_status, index=4 if len(cols_status) >= 5 else 0)
                except Exception as e:
                    st.error(f"Erro ao ler status: {e}")
            
            st.markdown("##### 🧑‍🤝‍🧑 3. Equipes de Apoio (Temporários)")
            with st.container(border=True):
                temp_bases_files = st.file_uploader("Suba a(s) planilha(s) de Apoio", type=["xlsx", "xls"], accept_multiple_files=True)
            
            df_bases_temp = pd.DataFrame()
            if temp_bases_files:
                try:
                    dfs_temp = []
                    for f in temp_bases_files:
                        df_t = ler_planilha_cached(f.getvalue())
                        df_t.columns = normalize_cols(df_t.columns)
                        if 'LEVANTADOR' not in df_t.columns:
                            for p_nome in ['NOME', 'TECNICO', 'EQUIPE']:
                                if p_nome in df_t.columns: df_t = df_t.rename(columns={p_nome: 'LEVANTADOR'}); break
                        dfs_temp.append(df_t)
                    df_bases_temp_full = pd.concat(dfs_temp, ignore_index=True)
                    
                    if 'LEVANTADOR' in df_bases_temp_full.columns:
                        opcoes_levs_temp = sorted([str(x) for x in df_bases_temp_full['LEVANTADOR'].dropna().unique().tolist() if str(x).upper().strip() != 'SEM LEVANTADOR'])
                        
                        levs_temp_selecionados = st.multiselect("Selecione as Equipes:", opcoes_levs_temp, default=opcoes_levs_temp)
                        
                        if levs_temp_selecionados:
                            df_bases_temp = df_bases_temp_full[df_bases_temp_full['LEVANTADOR'].isin(levs_temp_selecionados)].copy()
                            
                            if 'LATITUDE' in df_bases_temp.columns and 'LONGITUDE' in df_bases_temp.columns:
                                df_bases_temp['LATITUDE'] = pd.to_numeric(df_bases_temp['LATITUDE'].astype(str).str.replace(',', '.'), errors='coerce')
                                df_bases_temp['LONGITUDE'] = pd.to_numeric(df_bases_temp['LONGITUDE'].astype(str).str.replace(',', '.'), errors='coerce')
                            elif 'RESIDENCIA' in df_bases_temp.columns or 'MUNICIPIO' in df_bases_temp.columns:
                                col_ref_temp = 'RESIDENCIA' if 'RESIDENCIA' in df_bases_temp.columns else 'MUNICIPIO'
                                muns_unicos_temp = df_bases_temp[col_ref_temp].dropna().unique()
                                mapa_coords_temp = {}
                                with st.spinner("🌍 Buscando coordenadas bases temporárias..."):
                                    for mun in muns_unicos_temp:
                                        lat, lon = obter_coordenadas_municipio_cached(mun)
                                        mapa_coords_temp[mun] = (lat, lon)
                                df_bases_temp['LATITUDE'] = df_bases_temp[col_ref_temp].map(lambda x: mapa_coords_temp.get(x, (np.nan, np.nan))[0])
                                df_bases_temp['LONGITUDE'] = df_bases_temp[col_ref_temp].map(lambda x: mapa_coords_temp.get(x, (np.nan, np.nan))[1])
                            else:
                                df_bases_temp['LATITUDE'] = np.nan
                                df_bases_temp['LONGITUDE'] = np.nan
                                
                            df_bases_temp = df_bases_temp.dropna(subset=['LATITUDE', 'LONGITUDE'])
                            df_bases_temp['TIPO_EQUIPE'] = 'TEMPORARIA'
                except Exception as e:
                    st.error(f"Erro: {e}")

            if not task_files: 
                st.info("Aguardando upload de obras na Base Principal.")
                return

            try:
                dfs = []
                for f in task_files:
                    df_temp = ler_planilha_cached(f.getvalue())
                    if len(dfs) == 0: st.session_state.colunas_originais = df_temp.columns.tolist()
                    df_temp.columns = normalize_cols(df_temp.columns)
                    dfs.append(df_temp)
                df_tasks = pd.concat(dfs, ignore_index=True)
                total_obras_inicial = len(df_tasks)
                
                for c_nome in ['CONTA CONTRATO', 'INSTALACAO']:
                    if c_nome in df_tasks.columns:
                        df_tasks[c_nome] = df_tasks[c_nome].astype(str).str.replace(r'\.0$', '', regex=True).replace('nan', '-')
                        
            except Exception as e:
                st.error(f"Erro ao unificar planilhas: {e}"); return

            if not df_status_upload.empty and coluna_status_selecionada:
                df_tasks = atualizar_status_via_df(df_tasks, df_status_upload, coluna_status_selecionada)

        st.markdown("---")
        
        if 'LATITUDE' not in df_tasks.columns: df_tasks['LATITUDE'] = np.nan
        if 'LONGITUDE' not in df_tasks.columns: df_tasks['LONGITUDE'] = np.nan

        df_tasks['LATITUDE'] = pd.to_numeric(df_tasks['LATITUDE'].astype(str).str.replace(',', '.'), errors='coerce')
        df_tasks['LONGITUDE'] = pd.to_numeric(df_tasks['LONGITUDE'].astype(str).str.replace(',', '.'), errors='coerce')

        c_ex1, c_ex2 = st.columns(2)
        with st.expander("🛠️ 4. Configuração de Filtros Iniciais", expanded=True):
            c_filt1, c_filt2 = st.columns(2)
            
            if 'STATUS LIST' in df_tasks.columns:
                df_tasks['STATUS_LIMPO'] = df_tasks['STATUS LIST'].astype(str).str.strip().str.upper()
                status_unicos = sorted(df_tasks['STATUS_LIMPO'].unique().tolist())
                padroes_ativos = [s for s in status_unicos if s in STATUS_PADRAO]
                status_selecionados = c_filt1.multiselect("📌 Filtrar Status de Início:", options=status_unicos, default=padroes_ativos)
                if not status_selecionados: st.warning("Selecione um status."); return
                df_tasks = df_tasks[df_tasks['STATUS_LIMPO'].isin(status_selecionados)].drop(columns=['STATUS_LIMPO'])

            if 'TIPO NOTA' in df_tasks.columns:
                tipos_nota_unicos = sorted(df_tasks['TIPO NOTA'].astype(str).dropna().unique().tolist())
                tipos_selecionados = c_filt2.multiselect("🏷️ Filtrar TIPO DE NOTA (Todas as Obras):", tipos_nota_unicos, default=tipos_nota_unicos)
                if not tipos_selecionados: st.warning("Selecione um Tipo de Nota."); return
                df_tasks = df_tasks[df_tasks['TIPO NOTA'].astype(str).isin(tipos_selecionados)]

        erros_coords_mask = df_tasks['LATITUDE'].isna() | df_tasks['LONGITUDE'].isna() | (df_tasks['LATITUDE'] == 0.0) | (df_tasks['LONGITUDE'] == 0.0)
        qtd_erros_iniciais = erros_coords_mask.sum()
        
        if qtd_erros_iniciais > 0:
            col_end = 'ENDERECO' if 'ENDERECO' in df_tasks.columns else ('RUA' if 'RUA' in df_tasks.columns else None)
            col_mun = 'MUNICIPIO' if 'MUNICIPIO' in df_tasks.columns else ('CIDADE' if 'CIDADE' in df_tasks.columns else None)
            
            if col_end and col_mun:
                with st.status(f"🛰️ Resgatando {qtd_erros_iniciais} obras sem coordenadas via Satélite...", expanded=True) as status:
                    st.write("Conectando ao OpenStreetMap...")
                    my_bar = st.progress(0.0)
                    
                    df_erros = df_tasks[erros_coords_mask].copy()
                    df_ok = df_tasks[~erros_coords_mask].copy()
                    
                    lats, lons = [], []
                    for i, row in enumerate(df_erros.itertuples()):
                        end_val = getattr(row, col_end)
                        mun_val = getattr(row, col_mun)
                        lat, lon = geocode_endereco_nominatim(end_val, mun_val)
                        lats.append(lat)
                        lons.append(lon)
                        time.sleep(0.6) 
                        my_bar.progress((i + 1) / qtd_erros_iniciais)
                        
                    df_erros['LATITUDE'] = lats
                    df_erros['LONGITUDE'] = lons
                    my_bar.empty()
                    
                    ainda_com_erro = df_erros['LATITUDE'].isna() | df_erros['LONGITUDE'].isna() | (df_erros['LATITUDE'] == 0.0)
                    resgatadas = (~ainda_com_erro).sum()
                    if resgatadas > 0:
                        status.update(label=f"✅ {resgatadas} coordenadas recuperadas e incluídas no roteamento!", state="complete", expanded=False)
                    else:
                        status.update(label="Falha ao resgatar coordenadas. Verifique a grafia dos endereços.", state="error", expanded=False)
                    
                    df_tasks = pd.concat([df_ok, df_erros[~ainda_com_erro]])
                    erros_coords_mask = df_tasks['LATITUDE'].isna() | df_tasks['LONGITUDE'].isna() | (df_tasks['LATITUDE'] == 0.0)
                
        qtd_erros_coords_finais = erros_coords_mask.sum()
        df_tasks = df_tasks[~erros_coords_mask]
        
        erros_nome = 0
        for col_nome in ['NOME', 'NOME DO SOLICITANTE', 'CLIENTE']:
            if col_nome in df_tasks.columns:
                erros_nome += df_tasks[col_nome].isna().sum()
                df_tasks = df_tasks.dropna(subset=[col_nome])
                df_tasks = df_tasks[df_tasks[col_nome].astype(str).str.strip() != '']
                
        if 'STATUS SAP' in df_tasks.columns: df_tasks = df_tasks[~df_tasks['STATUS SAP'].astype(str).str.strip().str.upper().isin(['CANC', 'FINL'])]
        if 'TIPO NOTA' in df_tasks.columns: df_tasks['PRIORIDADE'] = df_tasks['TIPO NOTA'].apply(lambda x: 'Sim' if str(x).strip().upper() in TIPOS_PRIORITARIOS else 'Não')
        else: df_tasks['PRIORIDADE'] = 'Não'

        df_tasks, qtd_condensada = fundir_super_pontos(df_tasks, raio_metros=5)
        if qtd_condensada > 0:
            st.toast(f"✅ Inteligência condensou {qtd_condensada} obras repetidas no mesmo endereço em 'Super Pontos'.")

        st.markdown("#### 📊 Raio-X da Base de Dados Carregada")
        st.markdown(f"""
        <div class="profiling-box">
            <b>Análise Estrutural:</b> Das {total_obras_inicial} linhas encontradas, o sistema filtrou e aprovou <b style="color: #0D256C;">{len(df_tasks)} paradas únicas</b> para roteamento. <br>
            <i>(Omitidos: {qtd_erros_coords_finais} sem coordenadas resolvíveis | {erros_nome} sem nome de cliente | Restante fora do status padrão).</i>
        </div>
        """, unsafe_allow_html=True)

        if df_tasks.empty: return

        # === ALOCAÇÃO TERRITORIAL (MÓDULO DE ALTA VELOCIDADE) ===
        df_tasks_alocadas = pd.DataFrame()
        bases_principais_records = df_bases.to_dict('records') if not df_bases.empty else []
        bases_temporarias_records = df_bases_temp.to_dict('records') if not df_bases_temp.empty else []
        todas_bases_records = bases_principais_records + bases_temporarias_records
        
        if len(todas_bases_records) > 0:
            df_tasks['BASE_ATRIBUIDA'] = "NÃO ALOCADO"
            
            df_tasks['COORD_KEY'] = df_tasks['LATITUDE'].astype(str) + "_" + df_tasks['LONGITUDE'].astype(str)
            coords_com_prio = df_tasks[df_tasks['PRIORIDADE'] == 'Sim']['COORD_KEY'].unique()
            df_tasks['PRECISA_PRINCIPAL'] = df_tasks['COORD_KEY'].isin(coords_com_prio)
            
            df_prio_e_agregadas = df_tasks[df_tasks['PRECISA_PRINCIPAL']].copy()
            df_comum_puro = df_tasks[~df_tasks['PRECISA_PRINCIPAL']].copy()
            
            col_mun_name = 'MUNICIPIO' if 'MUNICIPIO' in df_tasks.columns else ('CIDADE' if 'CIDADE' in df_tasks.columns else None)
            if col_mun_name:
                df_prio_e_agregadas['MUN_LIMPO'] = normalizar_municipios(df_prio_e_agregadas[col_mun_name].fillna(''))
                df_comum_puro['MUN_LIMPO'] = normalizar_municipios(df_comum_puro[col_mun_name].fillna(''))
            else:
                df_prio_e_agregadas['MUN_LIMPO'] = ''
                df_comum_puro['MUN_LIMPO'] = ''
            
            mun_to_main = {}
            mun_to_all = {}
            for b in todas_bases_records:
                muns_str = str(b.get('MUNICIPIO', b.get('RESIDENCIA', '')))
                for m in muns_str.split(','):
                    m_limpo = normalizar_municipios(pd.Series([m])).iloc[0]
                    if m_limpo:
                        if m_limpo not in mun_to_all: mun_to_all[m_limpo] = []
                        if b['LEVANTADOR'] not in mun_to_all[m_limpo]: mun_to_all[m_limpo].append(b['LEVANTADOR'])
                        if b.get('TIPO_EQUIPE') == 'PRINCIPAL':
                            if m_limpo not in mun_to_main: mun_to_main[m_limpo] = []
                            if b['LEVANTADOR'] not in mun_to_main[m_limpo]: mun_to_main[m_limpo].append(b['LEVANTADOR'])

            base_counts = {b['LEVANTADOR']: 0 for b in todas_bases_records}
            
            # --- CORREÇÃO DE COTA (BUG DO STREAMLIT STATE) ---
            # Agora ele olha diretamente para a escolha da tela e não para o state anterior
            dias_multiplier = 6 if tipo_periodo == 'Semana' else 1
            
            if modo_limite == "Quantidade Fixa de Obras":
                obras_dia_est = obras_por_dia
            else:
                obras_dia_est = int(horas_por_dia / tempo_medio_obra)
                if obras_dia_est < 1: obras_dia_est = 1
                
            max_capacity = obras_dia_est * dias_multiplier * limite_periodos

            def assign_load_balanced(df_sub, allowed_bases, is_prio=False):
                if df_sub.empty or not allowed_bases: return pd.DataFrame(), df_sub.copy()
                df_sub = df_sub.sort_values(by=['LATITUDE', 'LONGITUDE'])
                assigned_rows = []
                unassigned_rows = []
                
                valid_bases_cache = {}
                
                for idx, row in df_sub.iterrows():
                    qtd_real = len(row.get('_ORIGINAL_ROWS', [1])) if isinstance(row.get('_ORIGINAL_ROWS'), list) else 1
                    lat, lon = row.get('LATITUDE'), row.get('LONGITUDE')
                    mun_str = str(row.get('MUN_LIMPO', ''))
                    
                    if mun_str not in valid_bases_cache:
                        if tipo_atribuicao == "Por Municípios Atendidos (Lê texto da planilha)":
                            valid_names = set(mun_to_main.get(mun_str, [])) if is_prio else set(mun_to_all.get(mun_str, []))
                            valid_bases_cache[mun_str] = [b for b in allowed_bases if b['LEVANTADOR'] in valid_names]
                        else:
                            valid_bases_cache[mun_str] = allowed_bases
                    
                    valid_bases = valid_bases_cache[mun_str]
                    
                    best_base = None
                    best_dist = float('inf')
                    
                    if pd.notna(lat) and pd.notna(lon):
                        for b in valid_bases:
                            b_name = b['LEVANTADOR']
                            if base_counts[b_name] + qtd_real <= max_capacity:
                                b_lat, b_lon = b.get('LATITUDE'), b.get('LONGITUDE')
                                if pd.notna(b_lat) and pd.notna(b_lon):
                                    d = haversine_scalar(lat, lon, float(b_lat), float(b_lon))
                                    
                                    if tipo_atribuicao == "Por Proximidade Geográfica das Coordenadas (Ignora texto)" and d > 100.0:
                                        continue 
                                        
                                    if d < best_dist:
                                        best_dist = d
                                        best_base = b_name
                                    
                    if best_base:
                        base_counts[best_base] += qtd_real
                        row['BASE_ATRIBUIDA'] = best_base
                        assigned_rows.append(row)
                    else:
                        row['BASE_ATRIBUIDA'] = "NÃO ALOCADO"
                        unassigned_rows.append(row)
                        
                return pd.DataFrame(assigned_rows), pd.DataFrame(unassigned_rows)

            df_prio_alocadas, df_prio_restante = assign_load_balanced(df_prio_e_agregadas, bases_principais_records, is_prio=True)
            df_comum_alocadas_princ, df_comum_restante = assign_load_balanced(df_comum_puro, bases_principais_records, is_prio=False)
            
            if not df_comum_restante.empty and bases_temporarias_records:
                df_comum_alocadas_temp, df_comum_restante_final = assign_load_balanced(df_comum_restante, bases_temporarias_records, is_prio=False)
            else:
                df_comum_alocadas_temp = pd.DataFrame()
                df_comum_restante_final = df_comum_restante
                
            df_tasks = pd.concat([df_prio_alocadas, df_prio_restante, df_comum_alocadas_princ, df_comum_alocadas_temp, df_comum_restante_final])
            df_tasks = df_tasks.drop(columns=['COORD_KEY', 'PRECISA_PRINCIPAL', 'MUN_LIMPO'], errors='ignore')
            
            df_unallocated = df_tasks[df_tasks['BASE_ATRIBUIDA'] == "NÃO ALOCADO"]
            df_tasks_alocadas = df_tasks[df_tasks['BASE_ATRIBUIDA'] != "NÃO ALOCADO"].copy()

            tot_unallocated = sum(len(r['_ORIGINAL_ROWS']) if isinstance(r.get('_ORIGINAL_ROWS'), list) else 1 for _, r in df_unallocated.iterrows())
            st.session_state.tot_obras_nao_alocadas = tot_unallocated

            if df_tasks_alocadas.empty: 
                st.error("Nenhuma obra encontrou equipes com cobertura geográfica ou com limite diário disponível.")
                if tot_unallocated > 0: st.warning(f"⚠️ {tot_unallocated} obras ficaram de fora e não puderam ser roteirizadas. Aumente o Limite de Semanas/Dias na configuração ou adicione novos Levantadores na planilha.")
                return
                
            bases_records = todas_bases_records 

        # Menu de exportação final
        with st.expander("🛠️ 5. Configuração de Saída", expanded=True):
            todas_cols = df_tasks_alocadas.columns.tolist()
            cols_desejadas = ['PROTOCOLO', 'CONTA CONTRATO', 'INSTALACAO', 'NOME', 'ENDEREÇO', 'MUNICIPIO', 'LATITUDE', 'LONGITUDE', 'TIPO NOTA']
            cols_padrao = [c for c in normalize_cols(cols_desejadas) if c in todas_cols]
            colunas_exibir = st.multiselect("Colunas Visíveis nos Cartões (KML/Mapa)", todas_cols, default=cols_padrao)
            st.info("⚡ **Deduplicação Ativa:** Obras num raio de 5 metros foram transformadas em Super Pontos para otimização.")
            col_prioridade = "TIPO NOTA"

        if st.button("🚀 Iniciar Motor de Roteirização (OR-Tools)", type="primary", use_container_width=True):
            if df_tasks_alocadas.empty: st.error("Selecione equipes válidas."); return
            if 'TIPO NOTA' in df_tasks_alocadas.columns and 'TIPO NOTA' not in colunas_exibir: colunas_exibir.append('TIPO NOTA')

            st.session_state.tarefas_alocadas_inicialmente = len(df_tasks_alocadas)
            st.session_state.bases_records = bases_records
            st.session_state.tipo_periodo = tipo_periodo
            st.session_state.colunas_exibir = colunas_exibir
            st.session_state.col_prioridade = col_prioridade
            
            st.session_state.config_financeira = {
                'custo_combustivel': custo_combustivel,
                'consumo_veiculo': consumo_veiculo,
                'custo_hora_equipe': custo_hora_equipe
            }
            
            st.session_state.vrp_state = {
                'config': {
                    'modo_limite': modo_limite, 'velocidade_media_kmh': velocidade_media_kmh,
                    'tempo_medio_obra': tempo_medio_obra, 'horas_por_dia': horas_por_dia, 'limite_km_diario': limite_km_diario,
                    'obras_por_dia': obras_por_dia, 'tipo_periodo': tipo_periodo, 'limite_periodos': limite_periodos
                },
                'b_names': list(set([b['LEVANTADOR'] for b in bases_records])),
                'b_idx': 0, 'unvisited': df_tasks_alocadas.copy(), 'routed_data': [],
            }
            st.session_state.vrp_status = "RUNNING"
            tentar_rerun()

if __name__ == "__main__":
    view_roteirizador()
